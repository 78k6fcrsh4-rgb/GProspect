"""
output/exporter.py
------------------
ResultExporter — exports formatted grant prospects to CSV and Excel.

Takes the list of formatted result dictionaries from ResultFormatter
and writes them to files that Deborah's Place staff can open
directly in Excel or Google Sheets.

Two export formats:
    1. CSV — simple, universally compatible, opens in any spreadsheet app
    2. Excel — formatted .xlsx with column widths and header styling

The exporter also saves a run summary file after each prospecting run
so staff can track what the agent found over time.

Usage:
    from output.exporter import ResultExporter
    from agent.profile import OrgProfile

    profile  = OrgProfile.from_json("profiles/deborah_place.json")
    exporter = ResultExporter(profile)

    # Export to CSV
    path = exporter.export_csv(formatted_results)
    print(f"Results saved to: {path}")

    # Export to Excel
    path = exporter.export_excel(formatted_results)
    print(f"Results saved to: {path}")
"""

from __future__ import annotations

import csv
import os
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from agent.profile import OrgProfile


# ─────────────────────────────────────────────────────────────────────────────
# Column definitions
# Defines the order and display names of columns in the export.
# Edit this list to add, remove, or reorder columns.
# ─────────────────────────────────────────────────────────────────────────────

CSV_COLUMNS = [
    ("rank",                     "Rank"),
    ("score_final",              "Final Score"),
    ("score_composite",          "Composite Score"),
    ("funder_name",              "Funder Name"),
    ("program_name",             "Program Name"),
    ("application_deadline",     "Application Deadline"),
    ("days_remaining",           "Days Remaining"),
    ("award_range",              "Award Range"),
    ("award_min",                "Award Min"),
    ("award_max",                "Award Max"),
    ("next_action",              "Recommended Next Action"),
    ("is_prior_funder",          "Prior Funder"),
    ("geographic_focus",         "Geographic Focus"),
    ("eligibility_requirements", "Eligibility Requirements"),
    ("application_url",          "Application URL"),
    ("application_method",       "Application Method"),
    ("program_officer",          "Program Officer"),
    ("funder_website",           "Funder Website"),
    ("score_geographic",         "Score: Geographic Alignment"),
    ("reason_geographic",        "Reason: Geographic"),
    ("score_population",         "Score: Population Alignment"),
    ("reason_population",        "Reason: Population"),
    ("score_budget",             "Score: Budget Fit"),
    ("reason_budget",            "Reason: Budget"),
    ("score_timeline",           "Score: Timeline Feasibility"),
    ("reason_timeline",          "Reason: Timeline"),
    ("description",              "Description"),
    ("focus_areas",              "Focus Areas"),
    ("required_documents",       "Required Documents"),
    ("disqualifying_factors",    "Disqualifying Factors"),
    ("completeness_notes",       "Completeness Notes"),
    ("who_complete",             "Who Complete"),
    ("how_complete",             "How Complete"),
    ("source",                   "Data Source"),
    ("source_url",               "Source URL"),
    ("date_found",               "Date Found"),
    ("org_name",                 "Organization"),
]


class ResultExporter:
    """
    Exports formatted grant prospect results to CSV and Excel files.

    Creates an output directory structure organized by org and date
    so runs are easy to find and compare over time.

    Output directory structure:
        outputs/
            deborah_place/
                2026-05-19/
                    grant_prospects_2026-05-19_143022.csv
                    grant_prospects_2026-05-19_143022.xlsx
                    run_summary_2026-05-19_143022.txt
    """

    def __init__(
        self,
        profile:    OrgProfile,
        output_dir: str = "outputs"
    ) -> None:
        """
        Initialize the exporter with the org profile.

        Args:
            profile:    Loaded and validated OrgProfile.
            output_dir: Base directory for all output files.
                       Defaults to 'outputs' in the project root.
        """
        self.profile    = profile
        self.output_dir = Path(output_dir)
        self.run_time   = datetime.now()
        self.run_dir    = self._create_run_directory()

    def export_csv(
        self,
        formatted_results: list[dict],
        filename: Optional[str] = None
    ) -> str:
        """
        Exports formatted results to a CSV file.

        Args:
            formatted_results: List of formatted result dictionaries
                              from ResultFormatter.format_all()
            filename:         Optional custom filename. If not provided,
                             a timestamped filename is generated.

        Returns:
            Absolute path to the created CSV file as a string.
        """
        if not filename:
            timestamp = self.run_time.strftime("%Y-%m-%d_%H%M%S")
            filename  = f"grant_prospects_{timestamp}.csv"

        filepath = self.run_dir / filename

        # Get column headers and keys
        headers = [display for _, display in CSV_COLUMNS]
        keys    = [key for key, _ in CSV_COLUMNS]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames   = keys,
                extrasaction = "ignore"
            )

            # Write header row with display names
            writer.writerow(dict(zip(keys, headers)))

            # Write each result row
            for result in formatted_results:
                writer.writerow(result)

        print(f"[ResultExporter] CSV saved: {filepath}")
        print(f"[ResultExporter] {len(formatted_results)} opportunities exported")

        return str(filepath)

    def export_excel(
        self,
        formatted_results: list[dict],
        filename: Optional[str] = None
    ) -> Optional[str]:
        """
        Exports formatted results to a formatted Excel file.

        Requires openpyxl to be installed. If not available,
        falls back to CSV export gracefully.

        Args:
            formatted_results: List of formatted result dictionaries.
            filename:         Optional custom filename.

        Returns:
            Absolute path to the Excel file, or None if openpyxl
            is not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("[ResultExporter] openpyxl not installed — falling back to CSV")
            return self.export_csv(formatted_results, filename)

        if not filename:
            timestamp = self.run_time.strftime("%Y-%m-%d_%H%M%S")
            filename  = f"grant_prospects_{timestamp}.xlsx"

        filepath = self.run_dir / filename
        wb       = openpyxl.Workbook()
        ws       = wb.active
        ws.title = "Grant Prospects"

        # ── Header row styling ────────────────────────────────────────────────
        header_fill = PatternFill(
            start_color = "1C3C64",
            end_color   = "1C3C64",
            fill_type   = "solid"
        )
        header_font  = Font(color="FFFFFF", bold=True, size=10)
        header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        headers = [display for _, display in CSV_COLUMNS]
        keys    = [key for key, _ in CSV_COLUMNS]

        for col_idx, header in enumerate(headers, 1):
            cell              = ws.cell(row=1, column=col_idx, value=header)
            cell.fill         = header_fill
            cell.font         = header_font
            cell.alignment    = header_align

        # ── Data rows ─────────────────────────────────────────────────────────
        score_fill_high   = PatternFill(start_color="E1F5EE", end_color="E1F5EE", fill_type="solid")
        score_fill_medium = PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid")
        score_fill_low    = PatternFill(start_color="FCEBEB", end_color="FCEBEB", fill_type="solid")

        for row_idx, result in enumerate(formatted_results, 2):
            # Determine row color based on final score
            score = result.get("score_final", 0)
            try:
                score_val = float(score)
                if score_val >= 4.0:
                    row_fill = score_fill_high
                elif score_val >= 2.5:
                    row_fill = score_fill_medium
                else:
                    row_fill = score_fill_low
            except (TypeError, ValueError):
                row_fill = None

            for col_idx, key in enumerate(keys, 1):
                value         = result.get(key, "")
                cell          = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.alignment = Alignment(
                    horizontal = "left",
                    vertical   = "top",
                    wrap_text  = True
                )
                if row_fill:
                    cell.fill = row_fill

        # ── Column widths ─────────────────────────────────────────────────────
        column_widths = {
            1:  6,   # Rank
            2:  10,  # Final Score
            3:  12,  # Composite Score
            4:  30,  # Funder Name
            5:  40,  # Program Name
            6:  18,  # Deadline
            7:  12,  # Days Remaining
            8:  20,  # Award Range
            9:  12,  # Award Min
            10: 12,  # Award Max
            11: 45,  # Next Action
            12: 12,  # Prior Funder
            13: 20,  # Geographic Focus
            14: 50,  # Eligibility
            15: 40,  # Application URL
            16: 15,  # Application Method
            17: 25,  # Program Officer
            18: 35,  # Funder Website
        }

        for col_idx, width in column_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Set remaining columns to a default width
        for col_idx in range(len(column_widths) + 1, len(keys) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 35

        # Freeze the header row so it stays visible when scrolling
        ws.freeze_panes = "A2"

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws_summary        = wb.create_sheet("Run Summary")
        ws_summary["A1"]  = "Grant Prospecting Run Summary"
        ws_summary["A1"].font = Font(bold=True, size=14, color="1C3C64")

        summary_data = [
            ("Organization",        self.profile.org_name),
            ("Run Date",            self.run_time.strftime("%B %d, %Y")),
            ("Run Time",            self.run_time.strftime("%H:%M:%S")),
            ("Total Opportunities", len(formatted_results)),
            ("",                    ""),
            ("Top 5 Results",       ""),
        ]

        for i, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=i, column=2, value=str(value))

        for i, result in enumerate(formatted_results[:5], len(summary_data) + 4):
            ws_summary.cell(
                row    = i,
                column = 1,
                value  = f"{result['rank']}. {result['funder_name']}"
            )
            ws_summary.cell(
                row    = i,
                column = 2,
                value  = f"Score: {result['score_final']} | Deadline: {result['application_deadline']}"
            )

        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 60

        wb.save(filepath)
        print(f"[ResultExporter] Excel saved: {filepath}")
        print(f"[ResultExporter] {len(formatted_results)} opportunities exported")

        return str(filepath)

    def export_run_summary(
        self,
        formatted_results: list[dict],
        raw_count:          int = 0,
        filtered_count:     int = 0,
    ) -> str:
        """
        Saves a plain-text run summary file.

        Records what the agent found, what was filtered out, and
        the top results. Used for logging and audit trail.

        Args:
            formatted_results: Final formatted results.
            raw_count:         Total raw results before filtering.
            filtered_count:    Results that passed eligibility filter.

        Returns:
            Path to the summary file.
        """
        timestamp = self.run_time.strftime("%Y-%m-%d_%H%M%S")
        filename  = f"run_summary_{timestamp}.txt"
        filepath  = self.run_dir / filename

        lines = [
            "=" * 60,
            f"GRANT PROSPECTING RUN SUMMARY",
            f"Organization: {self.profile.org_name}",
            f"Date: {self.run_time.strftime('%B %d, %Y')}",
            f"Time: {self.run_time.strftime('%H:%M:%S')}",
            "=" * 60,
            f"",
            f"Raw results found:        {raw_count}",
            f"Passed eligibility filter: {filtered_count}",
            f"Final ranked results:      {len(formatted_results)}",
            f"",
            "=" * 60,
            "TOP RESULTS",
            "=" * 60,
        ]

        for result in formatted_results[:10]:
            lines.extend([
                f"",
                f"Rank {result['rank']}: {result['funder_name']}",
                f"  Program:  {result['program_name']}",
                f"  Score:    {result['score_final']}/5",
                f"  Deadline: {result['application_deadline']} ({result['days_remaining']} days)",
                f"  Award:    {result['award_range']}",
                f"  Action:   {result['next_action']}",
            ])

        lines.append("\n" + "=" * 60)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        print(f"[ResultExporter] Run summary saved: {filepath}")
        return str(filepath)

    # ── Private helpers ───────────────────────────────────────────────────────

    def _create_run_directory(self) -> Path:
        """
        Creates the output directory for this run.

        Structure: outputs/{org_short_name}/{date}/

        Returns:
            Path object for the run directory.
        """
        org_folder  = self.profile.org_short_name.lower().replace(" ", "_").replace("'", "")
        date_folder = self.run_time.strftime("%Y-%m-%d")
        run_dir     = self.output_dir / org_folder / date_folder

        run_dir.mkdir(parents=True, exist_ok=True)
        return run_dir